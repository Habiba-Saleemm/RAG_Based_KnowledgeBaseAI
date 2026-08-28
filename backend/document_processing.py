"""
document_processing.py

Turns an uploaded file's raw bytes into plain text, then splits that text
into smaller chunks so each chunk can be embedded and searched individually
(a full document is usually too long/unfocused to embed as one piece).
"""

import io
import re

from pypdf import PdfReader
from docx import Document as DocxDocument
from openpyxl import load_workbook


def extract_qa_pairs(text: str) -> list[dict]:
    """
    Detects Q&A pairs in either of these formats:

    Format 1:
        Q1. What is RAG?
        A1. RAG stands for Retrieval-Augmented Generation.

    Format 2:
        QUESTION: What is RAG?
        ANSWER: RAG stands for Retrieval-Augmented Generation.

    Returns a list of dictionaries containing question and answer.
    Returns [] if no valid Q&A pairs are found.
    """

    pattern = re.compile(
        r"(?:"
        r"Q\s*\d+\s*[.)]\s*|QUESTION:\s*"
        r")"
        r"(?P<question>.*?)\s*"
        r"(?:A\s*\d*\s*[.:)]\s*|ANSWER:\s*)"
        r"(?P<answer>.*?)"
        r"(?=\n\s*(?:Q\s*\d+\s*[.)]|QUESTION:)|\Z)",
        re.DOTALL | re.IGNORECASE,
    )

    pairs = []

    for match in pattern.finditer(text):
        question = " ".join(match.group("question").split())
        answer = " ".join(match.group("answer").split())

        if question and answer:
            pairs.append({
                "question": question,
                "answer": answer,
            })

    return pairs


def extract_text(filename: str, file_bytes: bytes) -> str:
    """Dispatches to the right extractor based on file extension."""

    lower_name = filename.lower()

    if lower_name.endswith(".pdf"):
        return _extract_pdf(file_bytes)

    if lower_name.endswith(".docx"):
        return _extract_docx(file_bytes)

    if lower_name.endswith(".xlsx"):
        return _extract_xlsx(file_bytes)

    if lower_name.endswith(".txt"):
        return file_bytes.decode("utf-8", errors="ignore")

    raise ValueError(f"Unsupported file type: {filename}")


def _extract_pdf(file_bytes: bytes) -> str:
    reader = PdfReader(io.BytesIO(file_bytes))

    pages = [
        page.extract_text() or ""
        for page in reader.pages
    ]

    return "\n".join(pages)


def _extract_docx(file_bytes: bytes) -> str:
    doc = DocxDocument(io.BytesIO(file_bytes))

    paragraphs = [
        p.text
        for p in doc.paragraphs
        if p.text.strip()
    ]

    return "\n".join(paragraphs)


def _extract_xlsx(file_bytes: bytes) -> str:
    workbook = load_workbook(
        io.BytesIO(file_bytes),
        data_only=True
    )

    lines = []

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):

            cells = [
                str(cell)
                for cell in row
                if cell is not None
            ]

            if cells:
                lines.append(" | ".join(cells))

    return "\n".join(lines)


def chunk_text(text: str, max_words: int = 150) -> list[str]:
    """
    Splits text into chunks of roughly max_words words each,
    breaking on paragraph boundaries where possible so chunks
    stay coherent rather than cutting mid-sentence.
    """

    paragraphs = [
        p.strip()
        for p in text.split("\n")
        if p.strip()
    ]

    chunks: list[str] = []
    current_words: list[str] = []

    for paragraph in paragraphs:

        paragraph_words = paragraph.split()

        if (
            len(current_words) + len(paragraph_words) > max_words
            and current_words
        ):
            chunks.append(" ".join(current_words))
            current_words = []

        current_words.extend(paragraph_words)

    if current_words:
        chunks.append(" ".join(current_words))

    return chunks